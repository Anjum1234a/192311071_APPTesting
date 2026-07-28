"""
192311071_APPTesting - Automated Excel Report Generator (.xlsx)
Generates executive summary dashboard, category breakdown, pass/fail metrics,
and detailed test case sheets for 325+ test cases using openpyxl.
"""

import os
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(test_results, output_path="E2E_Test_Report_APPTesting.xlsx"):
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles & Colors
    font_family = "Arial"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")

    card_bg_pass = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light Green
    card_txt_pass = Font(name=font_family, size=18, bold=True, color="166534")

    card_bg_fail = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light Red
    card_txt_fail = Font(name=font_family, size=18, bold=True, color="991B1B")

    card_bg_total = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # Light Blue
    card_txt_total = Font(name=font_family, size=18, bold=True, color="1E40AF")

    card_bg_deploy = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    card_txt_deploy = Font(name=font_family, size=14, bold=True, color="065F46")

    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    pass_font = Font(name=font_family, size=10, bold=True, color="15803D")

    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    fail_font = Font(name=font_family, size=10, bold=True, color="B91C1C")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    thick_box = Border(
        left=Side(style='medium', color='0F172A'),
        right=Side(style='medium', color='0F172A'),
        top=Side(style='medium', color='0F172A'),
        bottom=Side(style='medium', color='0F172A')
    )

    # Compute Statistics
    total_cases = len(test_results)
    passed_cases = sum(1 for r in test_results if r.status == "PASS")
    failed_cases = sum(1 for r in test_results if r.status == "FAIL")
    skipped_cases = sum(1 for r in test_results if r.status == "SKIP")
    pass_rate = round((passed_cases / total_cases * 100), 2) if total_cases > 0 else 100.0
    total_exec_time = round(sum(r.exec_time for r in test_results), 2)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Categories list
    categories = ["UI/UX Testing", "Functional E2E Testing", "Unit Testing", "Validation Testing", "Vulnerability & Security", "Deployment Readiness"]

    # =========================================================================
    # SHEET 1: EXECUTIVE SUMMARY
    # =========================================================================
    ws_sum = wb.create_sheet(title="Executive Summary")
    ws_sum.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_sum.merge_cells("B2:H2")
    ws_sum["B2"] = "SMART DENTAL & CLINIDENT - AUTOMATED E2E TEST REPORT"
    ws_sum["B2"].font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    ws_sum["B2"].fill = header_fill
    ws_sum["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[2].height = 40

    # Subtitle Info
    ws_sum.merge_cells("B3:H3")
    ws_sum["B3"] = f"Repository: Anjum1234a/192311071_APPTesting  |  Generated: {timestamp}  |  Environment: GitHub Actions / Headless Chrome"
    ws_sum["B3"].font = Font(name=font_family, size=9, italic=True, color="64748B")
    ws_sum["B3"].alignment = Alignment(horizontal="center", vertical="center")

    # KPI Metric Cards Row (Row 5 & 6)
    # Total
    ws_sum.merge_cells("B5:C5")
    ws_sum["B5"] = "TOTAL TEST CASES"
    ws_sum["B5"].font = Font(name=font_family, size=9, bold=True, color="1E40AF")
    ws_sum["B5"].alignment = Alignment(horizontal="center", vertical="center")

    ws_sum.merge_cells("B6:C6")
    ws_sum["B6"] = total_cases
    ws_sum["B6"].font = card_txt_total
    ws_sum["B6"].fill = card_bg_total
    ws_sum["B6"].alignment = Alignment(horizontal="center", vertical="center")

    # Passed
    ws_sum.merge_cells("D5:E5")
    ws_sum["D5"] = "PASSED"
    ws_sum["D5"].font = Font(name=font_family, size=9, bold=True, color="166534")
    ws_sum["D5"].alignment = Alignment(horizontal="center", vertical="center")

    ws_sum.merge_cells("D6:E6")
    ws_sum["D6"] = passed_cases
    ws_sum["D6"].font = card_txt_pass
    ws_sum["D6"].fill = card_bg_pass
    ws_sum["D6"].alignment = Alignment(horizontal="center", vertical="center")

    # Failed
    ws_sum.merge_cells("F5:F5")
    ws_sum["F5"] = "FAILED"
    ws_sum["F5"].font = Font(name=font_family, size=9, bold=True, color="991B1B")
    ws_sum["F5"].alignment = Alignment(horizontal="center", vertical="center")

    ws_sum["F6"] = failed_cases
    ws_sum["F6"].font = card_txt_fail
    ws_sum["F6"].fill = card_bg_fail
    ws_sum["F6"].alignment = Alignment(horizontal="center", vertical="center")

    # Pass Rate
    ws_sum.merge_cells("G5:H5")
    ws_sum["G5"] = "PASS RATE %"
    ws_sum["G5"].font = Font(name=font_family, size=9, bold=True, color="065F46")
    ws_sum["G5"].alignment = Alignment(horizontal="center", vertical="center")

    ws_sum.merge_cells("G6:H6")
    ws_sum["G6"] = f"{pass_rate}%"
    ws_sum["G6"].font = card_txt_pass
    ws_sum["G6"].fill = card_bg_pass
    ws_sum["G6"].alignment = Alignment(horizontal="center", vertical="center")

    ws_sum.row_dimensions[5].height = 20
    ws_sum.row_dimensions[6].height = 35

    # Deployable Status Banner (Row 8)
    ws_sum.merge_cells("B8:H8")
    deploy_status_str = "DEPLOYMENT STATUS: READY FOR PRODUCTION (PASSED ALL CRITICAL SUITES)" if failed_cases == 0 else "DEPLOYMENT STATUS: ACTION REQUIRED (INVESTIGATE FAILURES)"
    ws_sum["B8"] = deploy_status_str
    ws_sum["B8"].font = card_txt_deploy
    ws_sum["B8"].fill = card_bg_deploy
    ws_sum["B8"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[8].height = 30

    # Category Breakdown Table (Row 11)
    ws_sum["B10"] = "TEST BREAKDOWN BY CATEGORY"
    ws_sum["B10"].font = Font(name=font_family, size=12, bold=True, color="0F172A")

    cat_headers = ["Test Category", "Total Cases", "Passed", "Failed", "Pass Rate (%)", "Exec Time (s)", "Category Status"]
    for col_idx, h in enumerate(cat_headers, start=2):
        cell = ws_sum.cell(row=11, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[11].height = 26

    current_row = 12
    for cat in categories:
        cat_results = [r for r in test_results if r.category == cat]
        c_total = len(cat_results)
        c_pass = sum(1 for r in cat_results if r.status == "PASS")
        c_fail = sum(1 for r in cat_results if r.status == "FAIL")
        c_rate = round((c_pass / c_total * 100), 1) if c_total > 0 else 100.0
        c_time = round(sum(r.exec_time for r in cat_results), 2)
        c_status = "READY" if c_fail == 0 else "FAILED"

        ws_sum.cell(row=current_row, column=2, value=cat).font = Font(name=font_family, size=10, bold=True)
        ws_sum.cell(row=current_row, column=3, value=c_total).alignment = Alignment(horizontal="center")
        ws_sum.cell(row=current_row, column=4, value=c_pass).alignment = Alignment(horizontal="center")
        ws_sum.cell(row=current_row, column=5, value=c_fail).alignment = Alignment(horizontal="center")
        ws_sum.cell(row=current_row, column=6, value=f"{c_rate}%").alignment = Alignment(horizontal="center")
        ws_sum.cell(row=current_row, column=7, value=f"{c_time}s").alignment = Alignment(horizontal="center")

        st_cell = ws_sum.cell(row=current_row, column=8, value=c_status)
        st_cell.alignment = Alignment(horizontal="center")
        st_cell.font = pass_font if c_status == "READY" else fail_font
        st_cell.fill = pass_fill if c_status == "READY" else fail_fill

        for c in range(2, 9):
            ws_sum.cell(row=current_row, column=c).border = thin_border
        current_row += 1

    # Total Summary Row
    ws_sum.cell(row=current_row, column=2, value="TOTAL SUMMARY").font = Font(name=font_family, size=10, bold=True)
    ws_sum.cell(row=current_row, column=3, value=total_cases).font = Font(name=font_family, size=10, bold=True)
    ws_sum.cell(row=current_row, column=4, value=passed_cases).font = Font(name=font_family, size=10, bold=True)
    ws_sum.cell(row=current_row, column=5, value=failed_cases).font = Font(name=font_family, size=10, bold=True)
    ws_sum.cell(row=current_row, column=6, value=f"{pass_rate}%").font = Font(name=font_family, size=10, bold=True)
    ws_sum.cell(row=current_row, column=7, value=f"{total_exec_time}s").font = Font(name=font_family, size=10, bold=True)
    tot_st = ws_sum.cell(row=current_row, column=8, value="PASSED ALL" if failed_cases == 0 else "FAILURES DETECTED")
    tot_st.font = Font(name=font_family, size=10, bold=True, color="166534" if failed_cases == 0 else "991B1B")

    for c in range(2, 9):
        ws_sum.cell(row=current_row, column=c).border = thin_border
        ws_sum.cell(row=current_row, column=c).fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    # Adjust Column Widths on Executive Summary
    for col in ws_sum.columns:
        col_letter = get_column_letter(col[0].column)
        ws_sum.column_dimensions[col_letter].width = 22

    # =========================================================================
    # DETAIL SHEETS GENERATOR
    # =========================================================================
    detail_sheets = [
        ("All Test Cases (325+)", test_results),
        ("UI & UX Testing", [r for r in test_results if r.category == "UI/UX Testing"]),
        ("Functional E2E Testing", [r for r in test_results if r.category == "Functional E2E Testing"]),
        ("Unit Testing", [r for r in test_results if r.category == "Unit Testing"]),
        ("Validation Testing", [r for r in test_results if r.category == "Validation Testing"]),
        ("Vulnerability & Security", [r for r in test_results if r.category == "Vulnerability & Security"]),
        ("Deployment Readiness", [r for r in test_results if r.category == "Deployment Readiness"])
    ]

    detail_headers = ["Test Case ID", "Category", "Sub-Module", "Test Title", "Description / Steps", "Expected Result", "Status", "Exec Time (s)", "Risk Level", "Notes"]

    for sheet_title, sheet_data in detail_sheets:
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True

        # Header Row
        for col_idx, h in enumerate(detail_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # Data Rows
        for row_idx, r in enumerate(sheet_data, start=2):
            ws.cell(row=row_idx, column=1, value=r.tc_id).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value=r.category)
            ws.cell(row=row_idx, column=3, value=r.submodule)
            ws.cell(row=row_idx, column=4, value=r.title)
            ws.cell(row=row_idx, column=5, value=r.description)
            ws.cell(row=row_idx, column=6, value=r.expected)

            st_cell = ws.cell(row=row_idx, column=7, value=r.status)
            st_cell.alignment = Alignment(horizontal="center")
            if r.status == "PASS":
                st_cell.fill = pass_fill
                st_cell.font = pass_font
            else:
                st_cell.fill = fail_fill
                st_cell.font = fail_font

            ws.cell(row=row_idx, column=8, value=r.exec_time).alignment = Alignment(horizontal="right")

            risk_cell = ws.cell(row=row_idx, column=9, value=r.severity)
            risk_cell.alignment = Alignment(horizontal="center")
            if r.severity == "CRITICAL":
                risk_cell.font = Font(name=font_family, size=9, bold=True, color="991B1B")
            elif r.severity == "HIGH":
                risk_cell.font = Font(name=font_family, size=9, bold=True, color="C2410C")
            else:
                risk_cell.font = Font(name=font_family, size=9, color="334155")

            ws.cell(row=row_idx, column=10, value=r.notes if r.notes else "Verified clean")

            for c in range(1, 11):
                ws.cell(row=row_idx, column=c).border = thin_border

        # Adjust Column Widths
        widths = [16, 22, 22, 35, 45, 45, 12, 14, 14, 25]
        for idx, w in enumerate(widths, start=1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = w

    # Save Workbook
    wb.save(output_path)
    print(f"[SUCCESS] E2E Test Report generated successfully: {os.path.abspath(output_path)}")
    return output_path

if __name__ == "__main__":
    from test_suite import E2ETestSuite
    suite = E2ETestSuite()
    results = suite.run_all_tests()
    generate_excel_report(results)
